import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


def normalize(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


def convert_excel_to_json(input_path: Path, output_path: Path, sheet_name: str | None = None) -> None:
    workbook = load_workbook(filename=input_path, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]

    rows = []
    for row in worksheet.iter_rows(min_row=2, min_col=1, max_col=11, values_only=True):
        col_a = normalize(row[0])
        fase = normalize(row[5])
        domein = normalize(row[6])
        subdomein = normalize(row[7])
        col_i = normalize(row[8])
        col_j = normalize(row[9])
        col_k = normalize(row[10])

        if all(value is None for value in (col_a, fase, domein, subdomein, col_i, col_j, col_k)):
            continue

        rows.append(
            {
                "colA": col_a,
                "fase": fase,
                "domein": domein,
                "subdomein": subdomein,
                "colI": col_i,
                "colJ": col_j,
                "colK": col_k,
            }
        )

    output_path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Klaar: {len(rows)} records geschreven naar {output_path}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Zet Excel om naar JSON met kolom A en kolommen F-K.")
    parser.add_argument("--input", required=True, help="Pad naar input Excel-bestand.")
    parser.add_argument("--output", required=True, help="Pad naar output JSON-bestand.")
    parser.add_argument("--sheet", default=None, help="Optionele sheetnaam.")
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    convert_excel_to_json(Path(args.input), Path(args.output), args.sheet)
